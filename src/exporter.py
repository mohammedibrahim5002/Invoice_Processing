"""
src/exporter.py
===============
Takes extracted fields from parser.py and model.py
and writes formatted Excel or CSV output.

Two modes:
  1. Single invoice  -> one Excel file with fields + confidence
  2. Batch           -> one Excel file, one row per invoice

Usage:
    from src.exporter import export_single, export_batch, export_csv

    # Single invoice
    fields = {'date': '25/12/2018', 'total': '9.00', ...}
    export_single(fields, 'outputs/result.xlsx')

    # Batch
    results = [fields1, fields2, fields3]
    export_batch(results, 'outputs/batch.xlsx')

    # CSV
    export_csv(results, 'outputs/batch.csv')
"""

import csv
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ── Colours ───────────────────────────────────────────────────────────────────

DARK_BLUE   = '1F4E79'
MID_BLUE    = '2E75B6'
LIGHT_BLUE  = 'DEEAF1'
LIGHTER     = 'EBF3FA'
WHITE       = 'FFFFFF'
GREY_TEXT   = '555555'
GREEN       = '375623'
GREEN_BG    = 'E2EFDA'
ORANGE      = '843C0C'
ORANGE_BG   = 'FCE4D6'
BORDER_COL  = 'BFBFBF'


# ── Style helpers ─────────────────────────────────────────────────────────────

def _border(color=BORDER_COL):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _font(bold=False, color='000000', size=11, italic=False):
    return Font(bold=bold, color=color, size=size,
                italic=italic, name='Calibri')

def _align(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _cell(ws, row, col, value='', bold=False, bg=None,
          font_color='000000', size=11, h_align='left',
          border=True, italic=False, num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, color=font_color, size=size, italic=italic)
    c.alignment = _align(h=h_align)
    if bg:
        c.fill  = _fill(bg)
    if border:
        c.border = _border()
    if num_format:
        c.number_format = num_format
    return c


# ── Field config ──────────────────────────────────────────────────────────────
# Defines display order, labels, and column widths for the Excel output

FIELD_CONFIG = [
    # (field_key,          display_label,        col_width)
    ('source_file',        'Source File',         30),
    ('invoice_number',     'Invoice Number',      18),
    ('date',               'Date',                14),
    ('due_date',           'Due Date',            14),
    ('vendor_name',        'Vendor Name',         28),
    ('company',            'Company (CRF)',        28),
    ('address',            'Address (CRF)',        40),
    ('total',              'Total Amount',        14),
    ('tax',                'Tax Amount',          12),
    ('tax_rate',           'Tax Rate',            10),
    ('currency',           'Currency',             9),
    ('email',              'Email',               28),
    ('phone',              'Phone',               16),
    ('payment_terms',      'Payment Terms',       18),
    ('extraction_method',  'Method',              12),
    ('extracted_at',       'Extracted At',        18),
]

CONFIDENCE_FIELDS = [
    'date', 'total', 'tax', 'invoice_number',
    'vendor_name', 'company', 'address', 'currency',
    'email', 'phone',
]


# ── Data preparation ──────────────────────────────────────────────────────────

def prepare_row(fields: dict, source_file: str = '') -> dict:
    """
    Merge parser fields, CRF fields, and metadata into one flat dict.
    Also computes an overall confidence score.
    """
    row = {}

    # Source and metadata
    row['source_file']       = source_file
    row['extracted_at']      = datetime.now().strftime('%Y-%m-%d %H:%M')
    row['extraction_method'] = fields.get('_method', 'parser+crf')

    # All extractable fields
    for key, _, _ in FIELD_CONFIG:
        if key in ('source_file', 'extracted_at', 'extraction_method'):
            continue
        row[key] = fields.get(key, '')

    # Confidence scores
    parser_conf = fields.get('_confidence', {})
    crf_conf    = fields.get('_crf_confidence', {})
    conf = {}
    for f in CONFIDENCE_FIELDS:
        c = max(parser_conf.get(f, 0.0), crf_conf.get(f, 0.0))
        conf[f] = round(c, 2)
    row['_confidence'] = conf

    # Overall confidence = average of non-zero field confidences
    non_zero = [v for v in conf.values() if v > 0]
    row['_overall_confidence'] = round(
        sum(non_zero) / len(non_zero), 2
    ) if non_zero else 0.0

    return row


def confidence_color(score: float) -> str:
    """Return background hex color based on confidence score."""
    if score >= 0.85:  return GREEN_BG
    if score >= 0.60:  return LIGHTER
    if score >  0.0:   return ORANGE_BG
    return WHITE


# ── Single invoice export ─────────────────────────────────────────────────────

def export_single(
    fields      : dict,
    output_path : str,
    source_file : str = '',
) -> None:
    """
    Export one invoice's extracted fields to a formatted Excel file.
    Creates two sheets: Summary and Confidence.

    Args:
        fields      : merged dict from parser + CRF
        output_path : where to save the .xlsx file
        source_file : original filename for reference
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    row = prepare_row(fields, source_file)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Extracted Fields ─────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Extracted Fields'
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.row_dimensions[1].height     = 30

    # Header
    _cell(ws, 1, 1, 'Invoice OCR Extraction', bold=True,
          bg=DARK_BLUE, font_color=WHITE, size=13, border=False)
    _cell(ws, 1, 2, source_file or 'Unknown',
          bg=DARK_BLUE, font_color='BDD7EE', size=11, border=False)
    _cell(ws, 1, 3, '', bg=DARK_BLUE, border=False)

    # Column headers
    ws.row_dimensions[2].height = 22
    for col, label in enumerate(['Field', 'Value', 'Confidence'], 1):
        _cell(ws, 2, col, label, bold=True,
              bg=MID_BLUE, font_color=WHITE, h_align='center')

    # Data rows
    conf = row.get('_confidence', {})
    r    = 3
    for key, label, _ in FIELD_CONFIG:
        if key in ('source_file', 'extraction_method', 'extracted_at'):
            continue
        val  = row.get(key, '')
        c    = conf.get(key, 0.0)
        bg   = confidence_color(c) if val else WHITE

        _cell(ws, r, 1, label, bold=True, bg=LIGHT_BLUE, size=10)
        _cell(ws, r, 2, str(val) if val else '—',
              bg=bg, size=10, italic=(not val))
        _cell(ws, r, 3,
              f'{c:.0%}' if c > 0 else '—',
              bg=bg, size=10, h_align='center')
        ws.row_dimensions[r].height = 18
        r += 1

    # Overall confidence row
    overall = row.get('_overall_confidence', 0.0)
    _cell(ws, r, 1, 'Overall Confidence', bold=True,
          bg=DARK_BLUE, font_color=WHITE, size=10)
    _cell(ws, r, 2, '', bg=DARK_BLUE)
    _cell(ws, r, 3, f'{overall:.0%}', bold=True,
          bg=DARK_BLUE, font_color=WHITE, h_align='center')

    # Metadata rows
    r += 2
    _cell(ws, r, 1, 'Extraction Method', bold=True, bg=LIGHT_BLUE, size=10)
    _cell(ws, r, 2, row.get('extraction_method', ''), size=10)
    _cell(ws, r, 3, '', border=False)
    r += 1
    _cell(ws, r, 1, 'Extracted At', bold=True, bg=LIGHT_BLUE, size=10)
    _cell(ws, r, 2, row.get('extracted_at', ''), size=10)
    _cell(ws, r, 3, '', border=False)

    wb.save(output_path)
    print(f'Saved: {output_path}')


# ── Batch export ──────────────────────────────────────────────────────────────

def export_batch(
    results     : list,
    output_path : str,
    source_files: list = None,
) -> None:
    """
    Export multiple invoices to one Excel file.
    Creates two sheets: Summary (one row per invoice) and Details.

    Args:
        results      : list of field dicts from parser + CRF
        output_path  : where to save the .xlsx file
        source_files : list of original filenames (same order as results)
    """
    if source_files is None:
        source_files = [f'invoice_{i+1}' for i in range(len(results))]

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    rows = [prepare_row(f, s) for f, s in zip(results, source_files)]

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum        = wb.active
    ws_sum.title  = 'Summary'

    # Column headers
    ws_sum.row_dimensions[1].height = 28
    for ci, (_, label, width) in enumerate(FIELD_CONFIG, 1):
        _cell(ws_sum, 1, ci, label, bold=True,
              bg=DARK_BLUE, font_color=WHITE, h_align='center')
        ws_sum.column_dimensions[get_column_letter(ci)].width = width

    # Add confidence column headers
    conf_start = len(FIELD_CONFIG) + 1
    _cell(ws_sum, 1, conf_start, 'Overall Confidence', bold=True,
          bg=MID_BLUE, font_color=WHITE, h_align='center')
    ws_sum.column_dimensions[get_column_letter(conf_start)].width = 18

    ws_sum.freeze_panes = 'A2'
    ws_sum.auto_filter.ref = f'A1:{get_column_letter(conf_start)}1'

    # Data rows
    for ri, row in enumerate(rows, 2):
        bg = LIGHTER if ri % 2 == 0 else WHITE
        ws_sum.row_dimensions[ri].height = 18

        for ci, (key, _, _) in enumerate(FIELD_CONFIG, 1):
            val = row.get(key, '')
            _cell(ws_sum, ri, ci, str(val) if val else '',
                  bg=bg, size=10)

        # Overall confidence with colour coding
        overall = row.get('_overall_confidence', 0.0)
        conf_bg = confidence_color(overall)
        _cell(ws_sum, ri, conf_start,
              f'{overall:.0%}' if overall > 0 else '—',
              bg=conf_bg, size=10, h_align='center')

    # ── Sheet 2: Field Confidence ─────────────────────────────────────────────
    ws_conf       = wb.create_sheet('Confidence')
    ws_conf.row_dimensions[1].height = 28

    conf_cols = ['Source File'] + CONFIDENCE_FIELDS + ['Overall']
    for ci, label in enumerate(conf_cols, 1):
        _cell(ws_conf, 1, ci, label, bold=True,
              bg=DARK_BLUE, font_color=WHITE, h_align='center')
        ws_conf.column_dimensions[get_column_letter(ci)].width = (
            30 if ci == 1 else 14
        )

    ws_conf.freeze_panes = 'A2'

    for ri, row in enumerate(rows, 2):
        conf    = row.get('_confidence', {})
        overall = row.get('_overall_confidence', 0.0)

        _cell(ws_conf, ri, 1, row.get('source_file', ''), size=10)

        for ci, field in enumerate(CONFIDENCE_FIELDS, 2):
            c  = conf.get(field, 0.0)
            bg = confidence_color(c)
            _cell(ws_conf, ri, ci,
                  f'{c:.0%}' if c > 0 else '—',
                  bg=bg, size=10, h_align='center')

        # Overall
        bg = confidence_color(overall)
        _cell(ws_conf, ri, len(CONFIDENCE_FIELDS) + 2,
              f'{overall:.0%}' if overall > 0 else '—',
              bold=True, bg=bg, size=10, h_align='center')

    # ── Sheet 3: Stats ────────────────────────────────────────────────────────
    ws_stats       = wb.create_sheet('Stats')
    ws_stats.column_dimensions['A'].width = 28
    ws_stats.column_dimensions['B'].width = 16

    _cell(ws_stats, 1, 1, 'Batch Statistics', bold=True,
          bg=DARK_BLUE, font_color=WHITE, size=12, border=False)
    _cell(ws_stats, 1, 2, '', bg=DARK_BLUE, border=False)

    stats = [
        ('Total invoices processed', len(rows)),
        ('Avg overall confidence',
         f'{sum(r.get("_overall_confidence",0) for r in rows)/len(rows):.0%}' if rows else '0%'),
        ('Invoices with date',
         sum(1 for r in rows if r.get('date'))),
        ('Invoices with total',
         sum(1 for r in rows if r.get('total'))),
        ('Invoices with company',
         sum(1 for r in rows if r.get('company'))),
        ('Invoices with address',
         sum(1 for r in rows if r.get('address'))),
        ('Invoices with email',
         sum(1 for r in rows if r.get('email'))),
        ('Invoices with phone',
         sum(1 for r in rows if r.get('phone'))),
        ('Extracted at',
         datetime.now().strftime('%Y-%m-%d %H:%M')),
    ]

    for ri, (label, value) in enumerate(stats, 2):
        bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        _cell(ws_stats, ri, 1, label, bold=True, bg=bg, size=10)
        _cell(ws_stats, ri, 2, str(value), bg=bg, size=10, h_align='center')

    wb.save(output_path)
    print(f'Saved: {output_path}  ({len(rows)} invoices)')


# ── CSV export ────────────────────────────────────────────────────────────────

def export_csv(
    results     : list,
    output_path : str,
    source_files: list = None,
) -> None:
    """
    Export extracted fields to a flat CSV file.
    One row per invoice, no formatting.

    Args:
        results      : list of field dicts
        output_path  : where to save the .csv file
        source_files : list of original filenames
    """
    if source_files is None:
        source_files = [f'invoice_{i+1}' for i in range(len(results))]

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    rows = [prepare_row(f, s) for f, s in zip(results, source_files)]

    # CSV columns = all FIELD_CONFIG keys + confidence fields
    base_cols  = [key for key, _, _ in FIELD_CONFIG]
    conf_cols  = [f'conf_{f}' for f in CONFIDENCE_FIELDS]
    all_cols   = base_cols + ['overall_confidence'] + conf_cols

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=all_cols, extrasaction='ignore')
        writer.writeheader()

        for row in rows:
            conf    = row.get('_confidence', {})
            overall = row.get('_overall_confidence', 0.0)
            flat    = {k: row.get(k, '') for k in base_cols}
            flat['overall_confidence'] = f'{overall:.2f}'
            for field in CONFIDENCE_FIELDS:
                flat[f'conf_{field}'] = f'{conf.get(field, 0.0):.2f}'
            writer.writerow(flat)

    print(f'Saved: {output_path}  ({len(rows)} rows)')


# ── Quick test ────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import sys
    sys.path.insert(0, '.')
    from pathlib import Path

    Path('outputs').mkdir(exist_ok=True)

    # Mock extracted fields
    sample_fields = {
        'invoice_number' : 'TD01167104',
        'date'           : '25/12/2018',
        'due_date'       : '',
        'vendor_name'    : 'BOOK TA .K (TAMAN DAYA) SDN BHD',
        'company'        : 'BOOK TA .K (TAMAN DAYA) SDN BHD',
        'address'        : 'NO.53 55,57 & 59, JALAN SAGU 18, TAMAN DAYA, 81100 JOHOR BAHRU',
        'total'          : '9.00',
        'tax'            : '',
        'tax_rate'       : '',
        'currency'       : 'RM',
        'email'          : '',
        'phone'          : '',
        'payment_terms'  : '',
        '_confidence'    : {
            'date'    : 0.95,
            'total'   : 0.98,
            'company' : 0.72,
            'address' : 0.84,
            'currency': 0.90,
        },
        '_crf_confidence': {
            'company': 0.85,
            'address': 0.78,
        },
        '_method': 'parser+crf',
    }

    # Test single export
    export_single(sample_fields, 'outputs/test_single.xlsx',
                  source_file='X00016469612.jpg')
    print('Single export done.')

    # Test batch export
    batch = [sample_fields, sample_fields, sample_fields]
    files = ['invoice_001.jpg', 'invoice_002.jpg', 'invoice_003.jpg']
    export_batch(batch, 'outputs/test_batch.xlsx', source_files=files)
    print('Batch export done.')

    # Test CSV
    export_csv(batch, 'outputs/test_batch.csv', source_files=files)
    print('CSV export done.')

    print('\nAll exports successful. Check outputs/ folder.')
